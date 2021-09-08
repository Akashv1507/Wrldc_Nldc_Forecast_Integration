from typing import List
import pandas as pd 
from openpyxl import load_workbook

def appendWorkbook(workbookName:str, data:List, columnNo:int, rowStartInd:int ):
    """add data to column number of workbook, starting from rowStartInd

    Args:
        workbookName (str): workbook path
        data (List): data in list
        columnNo (int): coloumn no, 1,2 3....etc
        rowStartInd (int): starting of row
    """    
    workbook = load_workbook(filename=workbookName)
    sheet = workbook.active
    # adding each item of data list to cell defined by columnNo and row
    for i, value in enumerate(data, start=rowStartInd):
        sheet.cell(column=columnNo, row=i, value=value)

    workbook.save(filename=workbookName)

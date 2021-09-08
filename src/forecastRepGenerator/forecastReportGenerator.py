import datetime as dt
import pandas as pd 
from openpyxl import load_workbook
from src.fetchersForReport.daForecastFetch import DayaheadForecastedDemandFetchRepo
from src.fetchersForReport.intradayForecastFetch import IntradayForecastedDemandFetchRepo
from src.fetchersForReport.actualDemandFetch import ActualDemandFetchFromApi
from src.forecastRepGenerator.dataRestructuring import restructureData
from utils.appendIntermediateWorkbook import appendWorkbook
from utils.printUtils import printWithTs

def generateForecastReport(startDate:dt.datetime, endDate: dt.datetime, configDict:dict)->bool:
    """generate forecast report

    Args:
        startDate (dt.datetime): start date
        endDate (dt.datetime): end date
        configDict (dict): application config

    Returns:
        bool: return true if report generation successfull for all days
    """   

    countErr = 0 
    conStr: str = configDict["con_string_mis_warehouse"]
    #creating instances of classes
    obj_dayaheadForecastedDemandFetchRepo = DayaheadForecastedDemandFetchRepo(conStr)
    obj_IntradayForecastedDemandFetchRepo = IntradayForecastedDemandFetchRepo(conStr)
    obj_actualDemandFetchFromApi = ActualDemandFetchFromApi(configDict['tokenUrl'], configDict['apiBaseUrl'], configDict['clientId'], configDict['clientSecret'] )

    #iterating through each day
    currDate = startDate
    while currDate <= endDate:
        try:
            # reading raw workbook and saving intermediate workbook with name "intermediate_07_09_2021_WR.xlsx"
            rawForecastWorkbookName = configDict['rawExcelPath'] + "\\rawForecastWorkbook.xlsx"
            intermediateWorkbookName = configDict['rawExcelPath'] + "\\intermediate_" + dt.datetime.strftime(currDate, '%d_%m_%Y') + "_WR.xlsx"
            # loading raw workbook and saving intermediate
            workbook = load_workbook(filename=rawForecastWorkbookName)
            workbook.save(filename=intermediateWorkbookName)

            #fetch DA forecast into df and append data to intermediate excel
            printWithTs('DA forecast report generation started...', clr='magenta')
            daForecastDf = obj_dayaheadForecastedDemandFetchRepo.fetchForecastedDemand(currDate, currDate)
            restructredData = restructureData(daForecastDf, dataIdentifier= "daFor")
            for item in restructredData:
                appendWorkbook(workbookName=intermediateWorkbookName, data=item['data'], columnNo=item['columnNo'], rowStartInd=4)
            printWithTs('DA forecast report generation completed...', clr='green')

            #fetch Intraday forecast into df and append data to intermediate excel
            printWithTs('Intraday forecast report generation started...', clr='magenta')
            intradayForecastDf = obj_IntradayForecastedDemandFetchRepo.fetchIntradayForecastedDemand(currDate, currDate)
            restructredData = restructureData(intradayForecastDf, dataIdentifier= "intraFor")
            for item in restructredData:
                appendWorkbook(workbookName=intermediateWorkbookName, data=item['data'], columnNo=item['columnNo'], rowStartInd=4)
            printWithTs('Intraday forecast report generation completed...', clr='green')
            
            #fetch actual demand into df and append data to intermediate excel
            printWithTs('Actual demand report generation started...', clr='magenta')
            actualDemand = obj_actualDemandFetchFromApi.fetchDemandDataFromApi(currDate, currDate)
            restructredData = restructureData(actualDemand, dataIdentifier= "actDem")
            for item in restructredData:
                appendWorkbook(workbookName=intermediateWorkbookName, data=item['data'], columnNo=item['columnNo'], rowStartInd=4)
            printWithTs('Actual demand report generation completed...', clr='green')

            #generating final workbook name
            finalWorkbookName = configDict['localForecastCsvDumpPath'] + "\\" + dt.datetime.strftime(currDate, '%d_%m_%Y') + "_WR.xlsx"
            # at this point intermediate workbook will be our final workbook, loading intermediate workbook and saving as a final workbook
            workbook = load_workbook(filename=intermediateWorkbookName)
            workbook.save(filename=finalWorkbookName)
            printWithTs('Final Workbook save completed...', clr='green')
        
        except Exception as err :
            printWithTs(err, clr='red')
            countErr = countErr+1

        currDate += dt.timedelta(days=1)

    if countErr ==0:
        return True
    else:
        return False